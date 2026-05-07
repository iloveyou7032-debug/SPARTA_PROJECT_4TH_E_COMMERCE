"""
absa_sampling.py — ABSA 골든셋 층화 표본 추출 모듈 (v1.0 / 2026-05-04)

목표
----
preprocessed_absa.parquet → 1,000건 골든셋 추출 (수기 라벨링용 xlsx)

설계 원칙
---------
1. 브랜드 균등 (4 × 250 = 1,000건)
2. 1-3점 oversample (브랜드별 50%, 즉 125건) — 별점 편향(4-5점 92%) 보정
3. 최소 길이 필터 (content_len ≥ 20) — 정보량 확보
4. product_id diversity (가급적 다양한 상품 분산)
5. 0점(별점 결측 추정)은 풀에서 제외
6. xlsx 출력: [가이드라인] 시트 + [라벨링] 시트 (드롭다운 검증 포함)

ABSA 6-Aspect 속성 구조 (확정)
------------------------------
1. 핏/사이즈     (Fit & Sizing)         — 사이즈, 라이즈, 압박감, Y존
2. 소재/내구성   (Material & Durability) — 원단 특성, 보풀·변색·늘어짐
3. 기능성        (Functional Performance) — 신축성, 흡습속건, 통기성, 활동성  [X축]
4. 디자인        (Design & Aesthetics)   — 색상, 패턴, 실루엣
5. 브랜드/헤리티지 (Brand & Heritage)    — 충성도, 재구매, 정체성            [Y축]
6. 가격/가치     (Price & Value)         — 가성비, 프로모션
"""

from __future__ import annotations

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


# ─── 상수: ABSA 속성 정의 ─────────────────────────────────────
ASPECTS: list[str] = [
    '핏/사이즈',
    '소재/내구성',
    '기능성',
    '디자인',
    '브랜드/헤리티지',
    '가격/가치',
]

POLARITY_CODES: list[str] = ['P', 'N', 'U', 'X']
POLARITY_LEGEND: str = 'P=Positive  N=Negative  U=Neutral  X=Not Mentioned'

# 라벨링 시트에 노출할 메타 컬럼
# content_clean : 라벨러 주 독해 텍스트 — ABSA 모델 입력과 동일한 정제 문장
# content       : 원본 보조 참고용 (이모티콘/특수문자 포함)
META_COLS: list[str] = [
    'review_id', 'brand', 'cat1', 'cat2',
    'product_name', 'rating', 'content_len', 'content_clean', 'content',
]


# ════════════════════════════════════════════════════════════
# 1. 골든셋 샘플링
# ════════════════════════════════════════════════════════════

def sample_golden_set(
    parquet_path: str = './final_data/preprocessed_absa.parquet',
    n_per_brand: int = 250,
    low_rating_ratio: float = 0.5,
    min_len: int = 20,
    seed: int = 260504,
) -> pd.DataFrame:
    """
    브랜드 균등 + 1-3점 oversample 층화 표본 추출.

    Parameters
    ----------
    parquet_path     : ABSA 전처리 parquet 경로
    n_per_brand      : 브랜드당 추출 건수 (기본 250)
    low_rating_ratio : 1-3점 비율 (기본 0.5 → 브랜드당 125건)
    min_len          : 최소 content_len (기본 20)
    seed             : 재현성 시드

    Returns
    -------
    DataFrame : 메타 컬럼 + 라벨 빈 컬럼 포함 (총 n_per_brand × 4 행)
    """
    df = pd.read_parquet(parquet_path)

    # 0점(별점 결측) + 짧은 리뷰 제외
    df = df[(df['rating'] >= 1) & (df['content_len'] >= min_len)].copy()

    n_low  = int(round(n_per_brand * low_rating_ratio))
    n_high = n_per_brand - n_low

    sampled: list[pd.DataFrame] = []
    for brand in ['안다르', '젝시믹스', 'FILA', '룰루레몬']:
        sub = df[df['brand'] == brand]
        low  = sub[sub['rating'].between(1, 3)]
        high = sub[sub['rating'].between(4, 5)]

        # 풀 부족 시 가용 최대치 추출 후 부족분은 반대편에서 보충
        take_low  = min(n_low,  len(low))
        take_high = min(n_high, len(high))
        deficit   = (n_low - take_low) + (n_high - take_high)
        if deficit > 0:
            # 4-5점 풀에서 보충 (1-3점은 보통 더 희소)
            extra = high.drop(high.sample(take_high, random_state=seed).index)
            take_extra = min(deficit, len(extra))
            extra_sample = extra.sample(take_extra, random_state=seed + 1)
        else:
            extra_sample = pd.DataFrame()

        s_low  = low.sample(take_low,   random_state=seed) if take_low  else pd.DataFrame()
        s_high = high.sample(take_high, random_state=seed) if take_high else pd.DataFrame()
        sampled.append(pd.concat([s_low, s_high, extra_sample], ignore_index=True))

    out = pd.concat(sampled, ignore_index=True)
    out = out.sample(frac=1, random_state=seed).reset_index(drop=True)  # 셔플
    out.insert(0, 'sample_idx', range(1, len(out) + 1))

    # 메타 컬럼 정렬 + 라벨 빈 컬럼 부착
    keep = ['sample_idx'] + [c for c in META_COLS if c in out.columns]
    out = out[keep].copy()
    for asp in ASPECTS:
        out[asp] = ''
    out['메모'] = ''

    return out


# ════════════════════════════════════════════════════════════
# 2. 어노테이션 가이드라인 텍스트 (xlsx 시트1)
# ════════════════════════════════════════════════════════════

_GUIDELINE_BLOCKS: list[tuple[str, list[str]]] = [
    ('1. 라벨링 입력 형식', [
        '각 리뷰 1건에 대해 6개 속성 컬럼에 다음 4-class 중 하나 입력:',
        '  P = Positive  (긍정 언급)',
        '  N = Negative  (부정 언급)',
        '  U = Neutral   (사실 진술 / 양가 감정)',
        '  X = Not Mentioned (해당 속성 미언급, 기본값)',
        '※ 속성이 한 리뷰에서 여러 번 언급되면 종합 극성을 입력.',
    ]),
    ('2. 속성 정의', [
        '핏/사이즈        : 사이즈 정확도, 라이즈, 압박감, Y존, 라인감',
        '소재/내구성      : 원단 촉감·두께, 보풀·변색·늘어짐 등 시간 경과 결함',
        '기능성           : 신축성, 흡습속건, 통기성, 활동성, 보온/냉감     [X축 직결]',
        '디자인           : 색상, 패턴, 실루엣, 데일리 활용도',
        '브랜드/헤리티지  : 브랜드 충성도, 재구매·추천, 정체성("올드머니룩" 등)  [Y축 직결]',
        '가격/가치        : 가격, 가성비, 프로모션·할인',
    ]),
    ('3. 모호 표현 매핑 사전', [
        '"보풀이 일어난다"           → 소재/내구성 = N',
        '"Y존이 부각되지 않는다"     → 핏/사이즈 = P',
        '"엉덩이 라인이 예쁘다"      → 핏/사이즈 = P',
        '"올드머니룩 같다"           → 브랜드/헤리티지 = P',
        '"휠라스럽다 / 안다르 같다"  → 브랜드/헤리티지 = (문맥에 따라)',
        '"또 살 거예요 / 재구매 의사" → 브랜드/헤리티지 = P',
        '"땀이 빨리 마른다"          → 기능성 = P',
        '"운동할 때 편하다"          → 기능성 = P',
        '"일상에서 입기 좋다"        → 디자인 = P (스타일링이 아닌 데일리 활용도)',
        '"예쁘다 (단독)"             → 디자인 = P',
        '"신축성이 좋아 핏이 잘 잡힘"→ 기능성 = P AND 핏/사이즈 = P (다중 라벨)',
        '"가성비 좋다"               → 가격/가치 = P',
        '"비싸다"                    → 가격/가치 = N',
        '"좋아요 (단독, 맥락 無)"    → 모두 X (속성 추론 불가)',
    ]),
    ('4. 감성 극성 판별 기준', [
        '① 명시적 키워드 우선 : "좋다/별로다/만족/불편" 등이 있으면 그 극성 적용',
        '② 암묵 부정 (Hedge) : "그닥…", "기대만큼은 아니다" → N',
        '③ 반어/비꼼          : "참 잘도 안 늘어나네" → N (명백할 때만)',
        '④ 혼합 감정          :',
        '    - 동일 속성 내 P+N 동시 발생 → U',
        '    - 단, 한쪽이 80%+ 우세 시 우세 극성',
        '⑤ 정도 부사 무관     : "너무 좋다" / "약간 좋다" 모두 P (강도는 별도 분석)',
    ]),
    ('5. 교차 검증 프로세스 (목표 Cohen\'s κ ≥ 0.75)', [
        'Phase 1  파일럿 50건 — 라벨러 2명 독립 라벨링',
        'Phase 2  κ 측정 (속성×극성별 confusion matrix)',
        '         └ κ < 0.75 속성은 정의·경계선 보강 + 50건 재라벨링',
        'Phase 3  본 라벨링 1,000건 — 라벨러 2명 독립 작업',
        'Phase 4  불일치 케이스 → 3rd 라벨러(시니어) 중재 합의',
        '         └ P↔N 정반대 → 무조건 중재 / P↔U·N↔U → 다수결',
        'Phase 5  최종 골든셋 확정 + κ 리포트(전체 / 속성별)',
    ]),
    ('6. 작업 팁', [
        '- 라벨링 시트의 속성 컬럼은 드롭다운(P/N/U/X) 입력 검증이 적용되어 있다.',
        '- 기본값은 X(미언급). 언급된 속성만 P/N/U로 변경하면 된다.',
        '- 메모 컬럼에는 애매한 케이스의 판단 근거를 짧게 기록한다.',
        '- review_id는 추후 합의 라벨 매칭 키이므로 절대 수정하지 않는다.',
    ]),
]


# ════════════════════════════════════════════════════════════
# 3. xlsx 어노테이션 템플릿 작성
# ════════════════════════════════════════════════════════════

def build_annotation_template(
    df: pd.DataFrame,
    output_path: str = './final_data/absa_golden_set_1000.xlsx',
) -> str:
    """
    df → xlsx 2시트(가이드라인 + 라벨링) 저장. 속성 컬럼 드롭다운 검증 포함.
    """
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)

    wb = Workbook()
    _write_guideline_sheet(wb)
    _write_annotation_sheet(wb, df)

    wb.save(output_path)
    return output_path


def _write_guideline_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = '가이드라인'

    title_font   = Font(name='맑은 고딕', size=16, bold=True, color='FFFFFF')
    title_fill   = PatternFill('solid', fgColor='2F5496')
    header_font  = Font(name='맑은 고딕', size=12, bold=True, color='FFFFFF')
    header_fill  = PatternFill('solid', fgColor='4472C4')
    body_font    = Font(name='맑은 고딕', size=10)
    wrap         = Alignment(wrap_text=True, vertical='top')

    ws.cell(1, 1, 'ABSA 골든셋 라벨링 가이드라인 (v1.0 / 2026-05-04)')
    ws.cell(1, 1).font = title_font
    ws.cell(1, 1).fill = title_fill
    ws.cell(1, 1).alignment = Alignment(vertical='center', horizontal='center')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.row_dimensions[1].height = 28

    row = 3
    for header, lines in _GUIDELINE_BLOCKS:
        c = ws.cell(row, 1, header)
        c.font = header_font
        c.fill = header_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1
        for line in lines:
            cell = ws.cell(row, 1, line)
            cell.font = body_font
            cell.alignment = wrap
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            row += 1
        row += 1  # 블록 간 빈 줄

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30


def _write_annotation_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet('라벨링')

    header_font = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4472C4')
    label_fill  = PatternFill('solid', fgColor='FFF2CC')   # 라벨 입력 컬럼 강조
    border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF'),
    )

    # 1행: 안내 (상단 고정)
    ws.cell(1, 1, f'※ 입력값: {POLARITY_LEGEND}   |   기본값 X(미언급), 언급된 속성만 P/N/U로 변경')
    ws.cell(1, 1).font = Font(name='맑은 고딕', size=10, italic=True, color='C00000')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))

    # 2행: 헤더
    for j, col in enumerate(df.columns, 1):
        cell = ws.cell(2, j, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # 3행~: 데이터
    for i, row_vals in enumerate(df.itertuples(index=False), 3):
        for j, val in enumerate(row_vals, 1):
            cell = ws.cell(i, j, val if not pd.isna(val) else '')
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = border
            # 라벨 입력 컬럼 배경 강조
            if df.columns[j - 1] in ASPECTS:
                cell.fill = label_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # 컬럼 너비
    width_map = {
        'sample_idx': 8,     'review_id': 14,  'brand': 10,
        'cat1': 10,          'cat2': 12,       'product_name': 32,
        'rating': 8,         'content_len': 10,
        'content_clean': 70, 'content': 50,    '메모': 30,
    }
    for j, col in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(j)].width = (
            width_map.get(col, 14) if col not in ASPECTS else 14
        )

    # 행 높이 + 헤더 고정
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 24
    ws.freeze_panes = 'A3'

    # 데이터 유효성 검사 (P/N/U/X 드롭다운) — 속성 컬럼만
    dv = DataValidation(
        type='list',
        formula1=f'"{",".join(POLARITY_CODES)}"',
        allow_blank=True,
        showDropDown=False,  # False = 드롭다운 표시
        errorTitle='입력 오류',
        error='P / N / U / X 중 하나만 입력 가능합니다.',
    )
    ws.add_data_validation(dv)
    last_row = 2 + len(df)
    for j, col in enumerate(df.columns, 1):
        if col in ASPECTS:
            letter = get_column_letter(j)
            dv.add(f'{letter}3:{letter}{last_row}')


# ════════════════════════════════════════════════════════════
# 4. 마스터 함수 (편의용)
# ════════════════════════════════════════════════════════════

def build_golden_set(
    parquet_path: str = './final_data/preprocessed_absa.parquet',
    output_path:  str = './final_data/absa_golden_set_1000.xlsx',
    n_per_brand:  int = 250,
    low_rating_ratio: float = 0.5,
    min_len:      int = 20,
    seed:         int = 260504,
) -> dict:
    """
    파이프라인: parquet → 샘플링 → xlsx 출력. 통계 dict 반환.
    """
    df = sample_golden_set(
        parquet_path=parquet_path,
        n_per_brand=n_per_brand,
        low_rating_ratio=low_rating_ratio,
        min_len=min_len,
        seed=seed,
    )
    path = build_annotation_template(df, output_path)

    stats = {
        'total':        len(df),
        'output_path':  path,
        'brand_dist':   df['brand'].value_counts().to_dict(),
        'rating_dist':  df['rating'].value_counts().sort_index().to_dict(),
        'low_rating':   int(df['rating'].between(1, 3).sum()),
        'high_rating':  int(df['rating'].between(4, 5).sum()),
    }
    return stats

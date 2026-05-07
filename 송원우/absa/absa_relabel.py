"""
absa_relabel.py — v1 라벨링 → v2 변환 모듈 (2026-05-06)

목표
----
1. U 클래스 폐지 (3-class P/N/X로 변환)
2. 트리거 사전 매칭으로 v1 라벨과 충돌하는 케이스 자동 탐지 (재검토 후보)
3. 부정 톤 우세 룰 적용 (rating ≤ 2 + 강한 부정어 시 약한 P → X 강등)
4. 라벨러별 재검토 시트 자동 생성 (1-500은 L1, 501-1000은 L2 본인이 검토)

변환 정책 (Convert Policy)
--------------------------
[자동 변환]
  - U → X (정보량 미미)
  - 부정 톤 우세 + 약한 P 시그널 → X 강등 (브랜드/헤리티지 한정)

[재검토 후보 플래그]
  - v1 라벨이 P/N인데 트리거 사전 매칭 없음   → "P_NO_TRIGGER" / "N_NO_TRIGGER"
  - v1 라벨이 X인데 트리거 사전 매칭 있음     → "X_HAS_TRIGGER"
  - "편하다" 단독 P 라벨                       → "PYEONHADA_AMBIGUOUS"
  - 부정 톤 우세 + 약한 P                      → "NEG_TONE_WEAK_P"

[라벨러 검토 분담]
  - sample_idx 1-500   → 라벨러 1 검토 (본인 원본)
  - sample_idx 501-1000 → 라벨러 2 검토 (본인 원본)
"""

from __future__ import annotations

import re
import os
import pandas as pd
from typing import Iterable
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

GUIDELINE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'absa_guideline_v2.md')

# ─── 트리거 사전 (가이드라인 v2 §3 ~ §4 기반) ──────────────────
TRIGGERS_P: dict[str, list[str]] = {
    '핏/사이즈': [
        '정사이즈', '딱 맞', '잘 맞', '핏 좋', '핏 잘 잡', '핏감 좋',
        '라인 예쁘', '라인 살아', 'Y존 커버', '와이존 커버',
        '허리 잘 잡', '라이즈 좋', '잘 잡아', '안 흘러내', '안 처짐',
    ],
    '소재/내구성': [
        '부드럽', '보들', '매끄럽', '촉감 좋', '원단 좋',
        '두께 적당', '두께감 좋', '탄탄', '탄력 좋',
        '보풀 없', '변색 없', '안 늘어', '오래 입',
    ],
    '기능성': [
        '신축성', '잘 늘어', '흡습', '땀 빨리 마', '땀 안 차',
        '통기', '바람 잘 통', '활동성', '운동할 때 편', '움직이기 편',
        '보온', '따뜻', '안 춥', '냉감', '안 비치',
    ],
    '디자인': [
        '예쁘', '이쁘', '디자인 좋', '디자인 마음',
        '색깔 예쁘', '색감 예쁘', '컬러 예쁘', '색이 좋', '색 예쁘',
        '패턴 예쁘', '실루엣 예쁘',
        '데일리', '코디 잘', '어울', '여성스럽', '세련', '고급스럽',
        '양면 착용', '다른 색', '다른 컬러',
    ],
    '브랜드/헤리티지': [
        '또 살', '또 사', '다음에도 살', '다음에도 사', '다음에 살', '다음에 사',
        '재구매', '재주문', '한 번 더',
        '추천', '강추',
        '믿고 산', '믿고 사', '팬', '단골', '여기만',
        '답다', '스럽다', '감성', '느낌',
        '안다르답', '젝시믹스답', '휠라', '룰루레몬', '올드머니',
    ],
    '가격/가치': [
        '가성비', '이 가격', '저렴', '합리적', '가격 착', '가격 만족',
        '할인 좋', '세일',
    ],
}

TRIGGERS_N: dict[str, list[str]] = {
    '핏/사이즈': [
        '사이즈 작', '사이즈 큼', '작아요', '커요', '한치수',
        '조임', '끼임', '답답', 'Y존 부각', '와이존 부각',
        '허리 말림', '흘러내', '처짐', '늘어짐', '라인 이상',
    ],
    '소재/내구성': [
        '보풀', '늘어남', '변색', '후줄근',
        '까실', '따끔', '비침', '비쳐', '비치는',
        '후진 원단', '원단 별로', '얇', '거칠', '까칠',
        '세탁 후 변', '빨면',
    ],
    '기능성': [
        '신축성 부족', '안 늘어남', '빳빳', '뻣뻣',
        '땀 안 마', '땀 차',
        '운동할 때 불편', '활동 제한', '움직이기 불편',
        '춥다', '안 따뜻', '너무 덥',
        '기능 떨어', '기능 별로', '운동복 같지 않',
    ],
    '디자인': [
        '안 예쁘', '디자인 별로', '색감 별로', '색 별로',
        '패턴 별로', '실루엣 이상', '사진과 다름', '사진이랑 다',
        '촌스럽', '안 어울', '너무 화려', '너무 밋밋',
    ],
    '브랜드/헤리티지': [
        '다신 안', '다시는 안', '실망', '기대 이하', '비추',
        '답지 못', '브랜드 가치 없', '환불', '반품',
    ],
    '가격/가치': [
        '비싸', '너무 비쌈', '가격 대비 별로', '가격이 부담',
        '이 가격에 이걸', '가격 거품',
    ],
}

# 부정 톤 우세 강한 부정어
NEG_TONE_WORDS = re.compile(
    r'별로|후줄근|실망|기대\s?이하|최악|환불|반품|비추|짜증|후진|망함|쓰레기|짜증나'
)

# 약한 P 시그널 (브랜드/헤리티지에서 부정 톤 우세 시 X 강등 대상)
WEAK_BRAND_P = re.compile(
    r'또\s?사|다음에도\s?사|추가\s?구매|재구매|다음에\s?사'
)

# "편하다" 단독 검출 — 맥락 키워드 ±5어절 내 부재 시 ambiguous
PYEONHADA = re.compile(r'편하?[다요시아]|편함')
PYEONHADA_CONTEXTS = re.compile(
    r'운동|활동|움직|착용감|허리|사이즈|원단|감촉|기능'
)

ASPECTS = ['핏/사이즈', '소재/내구성', '기능성', '디자인', '브랜드/헤리티지', '가격/가치']


# ════════════════════════════════════════════════════════════
# 1. 매칭 유틸
# ════════════════════════════════════════════════════════════

def _has_any(text: str, keywords: Iterable[str]) -> bool:
    return any(k in text for k in keywords)


def detect_triggers(text: str) -> dict[str, str | None]:
    """
    텍스트 1건 → 속성별 트리거 매칭 결과
    Returns: {속성: 'P' | 'N' | None}  (둘 다 매칭 시 P 우선)
    """
    text = text or ''
    out = {}
    for asp in ASPECTS:
        has_p = _has_any(text, TRIGGERS_P[asp])
        has_n = _has_any(text, TRIGGERS_N[asp])
        if has_p and not has_n:
            out[asp] = 'P'
        elif has_n and not has_p:
            out[asp] = 'N'
        elif has_p and has_n:
            out[asp] = 'P'   # 양쪽 매칭 시 P 우선 (라벨러 검토 권장)
        else:
            out[asp] = None
    return out


def is_neg_tone(rating: float, text: str) -> bool:
    """부정 톤 우세 룰 (가이드라인 v2 §4.2)"""
    if pd.isna(rating) or rating > 2:
        return False
    return bool(NEG_TONE_WORDS.search(text or ''))


def is_pyeonhada_ambiguous(text: str) -> bool:
    """"편하다" 단독 (맥락 키워드 부재) 여부"""
    text = text or ''
    if not PYEONHADA.search(text):
        return False
    return not PYEONHADA_CONTEXTS.search(text)


# ════════════════════════════════════════════════════════════
# 2. v1 → v2 변환 + 재검토 플래그 부착
# ════════════════════════════════════════════════════════════

def convert_v1_to_v2(
    done_path: str = './final_data/absa_golden_set_1000_done.xlsx',
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    v1 라벨 데이터 → v2 변환 + 재검토 후보 플래그
    Returns: (df_v2, df_review_candidates)
    """
    df = pd.read_excel(done_path, sheet_name='라벨링', header=1)

    # 정규화 (NaN·기타 → X, 대소문자 통일)
    for col in ASPECTS:
        df[col] = df[col].fillna('X').astype(str).str.upper().str.strip()
        df.loc[~df[col].isin({'P', 'N', 'U', 'X'}), col] = 'X'

    df_v2 = df.copy()
    flags: list[dict] = []   # 재검토 후보 누적

    for i, row in df.iterrows():
        text = str(row.get('content_clean', '') or '')
        rating = row.get('rating')
        triggers = detect_triggers(text)
        neg_tone = is_neg_tone(rating, text)
        pyeon_ambig = is_pyeonhada_ambiguous(text)

        for asp in ASPECTS:
            v1 = row[asp]
            new_v = v1
            flag_reasons = []

            # ① U → X 자동 변환
            if v1 == 'U':
                new_v = 'X'
                flag_reasons.append('U→X (자동변환)')

            # ② 부정 톤 우세 + 약한 P (브랜드/헤리티지) → X 강등 후보
            if asp == '브랜드/헤리티지' and v1 == 'P' and neg_tone:
                if WEAK_BRAND_P.search(text):
                    flag_reasons.append('NEG_TONE_WEAK_P (강등 검토)')

            # ③ "편하다" 단독 P 라벨 → 모호 (4장 1절)
            if v1 == 'P' and pyeon_ambig and asp in ('핏/사이즈', '기능성', '소재/내구성'):
                flag_reasons.append('PYEONHADA_AMBIGUOUS (편하다 단독)')

            # ④ v1 P/N인데 트리거 매칭 없음 → 라벨 근거 약함
            if v1 in ('P', 'N') and triggers[asp] is None:
                flag_reasons.append(f'{v1}_NO_TRIGGER (트리거 매칭 없음)')

            # ⑤ v1 X인데 트리거 P/N 매칭 → 누락 의심
            if v1 == 'X' and triggers[asp] is not None:
                flag_reasons.append(f'X_HAS_TRIGGER ({triggers[asp]} 매칭)')

            # ⑥ v1 P/N과 트리거 매칭 결과 충돌 (P→N 또는 N→P)
            if v1 in ('P', 'N') and triggers[asp] is not None and v1 != triggers[asp]:
                flag_reasons.append(f'{v1}_vs_TRIGGER_{triggers[asp]} (충돌)')

            df_v2.at[i, asp] = new_v

            if flag_reasons:
                flags.append({
                    'sample_idx':    row['sample_idx'],
                    '라벨러':         '라벨러1' if row['sample_idx'] <= 500 else '라벨러2',
                    'review_id':     row.get('review_id'),
                    'brand':         row.get('brand'),
                    'rating':        row.get('rating'),
                    'content_clean': text[:200],
                    '속성':          asp,
                    'v1_라벨':       v1,
                    '자동변환':      new_v if new_v != v1 else '',
                    '트리거매칭':    triggers[asp] or '',
                    '플래그':        ' | '.join(flag_reasons),
                    '최종라벨':      '',   # 라벨러 입력 컬럼
                    '메모':          '',
                })

    df_review = pd.DataFrame(flags)
    return df_v2, df_review


# ════════════════════════════════════════════════════════════
# 3. 라벨러별 검토 시트 분리 저장 (P1 우선 / P2 보조)
# ════════════════════════════════════════════════════════════

# 우선순위 정의
# P1 (즉시 검토): 라벨러 v1 ↔ 트리거 사전 결과가 정면 충돌하거나 모호 룰 적용 대상
P1_FLAGS = {
    'X_HAS_TRIGGER',           # v1=X인데 트리거 매칭 → 누락 가능
    'P_vs_TRIGGER_N',          # v1=P, 트리거=N → 정반대
    'N_vs_TRIGGER_P',          # v1=N, 트리거=P → 정반대
    'PYEONHADA_AMBIGUOUS',     # "편하다" 단독 P → 모호
    'NEG_TONE_WEAK_P',         # 부정 톤 우세 + 약한 P
}

# P2 (선택 검토): 트리거 매칭 부재 (사전 한계일 가능성 높음 — 스폿 체크용)
P2_FLAGS = {'P_NO_TRIGGER', 'N_NO_TRIGGER'}


def _classify_priority(flag_str: str) -> str:
    """플래그 문자열에서 P1 우선 케이스 우선 분류"""
    flag_types = {f.split(' (')[0] for f in flag_str.split(' | ')}
    if flag_types & P1_FLAGS:
        return 'P1'
    if flag_types & P2_FLAGS:
        return 'P2'
    return 'P3'   # U→X 자동변환만 적용된 경우


def _read_guideline_lines() -> list[str]:
    """absa_guideline_v2.md 파일을 라인 리스트로 로드"""
    if not os.path.exists(GUIDELINE_PATH):
        return [f'[경고] 가이드라인 파일 없음: {GUIDELINE_PATH}']
    with open(GUIDELINE_PATH, encoding='utf-8') as f:
        return [line.rstrip() for line in f.readlines()]


def _write_guideline_sheet(wb: Workbook, sheet_name: str = '가이드라인') -> None:
    """
    absa_guideline_v2.md 내용을 워크북의 새 시트로 렌더링.
    헤더(#/##/###), 테이블(|...|), 코드블록(```), 리스트(-) 형식 인식.
    """
    ws = wb.create_sheet(sheet_name, 0)   # 첫 번째 시트로 삽입

    # 스타일 정의
    h1_font = Font(name='맑은 고딕', size=18, bold=True, color='FFFFFF')
    h1_fill = PatternFill('solid', fgColor='1F3864')
    h2_font = Font(name='맑은 고딕', size=14, bold=True, color='FFFFFF')
    h2_fill = PatternFill('solid', fgColor='2F5496')
    h3_font = Font(name='맑은 고딕', size=12, bold=True, color='1F3864')
    h3_fill = PatternFill('solid', fgColor='D9E1F2')
    table_header_font = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
    table_header_fill = PatternFill('solid', fgColor='4472C4')
    table_cell_font   = Font(name='맑은 고딕', size=10)
    table_cell_fill   = PatternFill('solid', fgColor='F2F2F2')
    body_font         = Font(name='맑은 고딕', size=10)
    code_font         = Font(name='Menlo', size=9, color='2F5496')
    code_fill         = PatternFill('solid', fgColor='F8F8F8')
    bold_font         = Font(name='맑은 고딕', size=10, bold=True)
    wrap = Alignment(wrap_text=True, vertical='top')
    border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF'),
    )

    lines = _read_guideline_lines()
    row = 1
    in_code_block = False
    in_table = False
    table_buffer: list[list[str]] = []

    def flush_table():
        nonlocal row, table_buffer, in_table
        if not table_buffer:
            return
        # 첫 줄 = 헤더, 둘째 줄 = 구분선(생략), 나머지 = 데이터
        header = table_buffer[0]
        body = [r for r in table_buffer[2:] if not all(re.fullmatch(r'-+', c.strip() or '-') for c in r)]
        n_cols = len(header)
        for j, cell_val in enumerate(header, 1):
            c = ws.cell(row, j, cell_val)
            c.font = table_header_font
            c.fill = table_header_fill
            c.alignment = wrap
            c.border = border
        row += 1
        for body_row in body:
            for j, cell_val in enumerate(body_row, 1):
                c = ws.cell(row, j, cell_val)
                c.font = table_cell_font
                c.fill = table_cell_fill
                c.alignment = wrap
                c.border = border
            row += 1
        table_buffer = []
        in_table = False
        row += 1   # 테이블 후 빈 줄

    for raw in lines:
        # 코드블록 토글
        if raw.strip().startswith('```'):
            if in_table: flush_table()
            in_code_block = not in_code_block
            continue

        # 테이블 라인 처리
        if raw.strip().startswith('|') and raw.strip().endswith('|') and not in_code_block:
            cells = [c.strip() for c in raw.strip('|').split('|')]
            table_buffer.append(cells)
            in_table = True
            continue
        elif in_table:
            flush_table()

        # 코드블록 내부
        if in_code_block:
            c = ws.cell(row, 1, raw)
            c.font = code_font
            c.fill = code_fill
            c.alignment = Alignment(wrap_text=True, vertical='top')
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            row += 1
            continue

        # 헤더 처리
        if raw.startswith('# '):
            c = ws.cell(row, 1, raw[2:])
            c.font, c.fill = h1_font, h1_fill
            c.alignment = Alignment(vertical='center', horizontal='left', indent=1)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.row_dimensions[row].height = 32
            row += 2
        elif raw.startswith('## '):
            c = ws.cell(row, 1, raw[3:])
            c.font, c.fill = h2_font, h2_fill
            c.alignment = Alignment(vertical='center', horizontal='left', indent=1)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.row_dimensions[row].height = 26
            row += 1
        elif raw.startswith('### '):
            c = ws.cell(row, 1, raw[4:])
            c.font, c.fill = h3_font, h3_fill
            c.alignment = Alignment(vertical='center', horizontal='left', indent=1)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.row_dimensions[row].height = 22
            row += 1
        elif raw.startswith('#### '):
            c = ws.cell(row, 1, raw[5:])
            c.font = bold_font
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            row += 1
        elif raw.startswith('---'):
            # 구분선 → 빈 줄로 처리
            row += 1
        elif raw.strip() == '':
            row += 1
        else:
            c = ws.cell(row, 1, raw)
            c.font = body_font
            c.alignment = Alignment(wrap_text=True, vertical='top')
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            row += 1

    flush_table()  # 마지막 테이블 잔여분

    # 컬럼 너비
    widths = [22, 22, 22, 22, 22, 22]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w


def _write_review_sheet(wb: Workbook, df: pd.DataFrame, sheet_name: str = '검토') -> None:
    """검토 데이터프레임을 워크북에 시트로 추가 + 최종라벨 컬럼 P/N/X 드롭다운 검증"""
    ws = wb.create_sheet(sheet_name)

    header_font = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4472C4')
    flag_fill   = PatternFill('solid', fgColor='FCE4D6')   # 플래그 컬럼 강조
    label_fill  = PatternFill('solid', fgColor='FFF2CC')   # 최종라벨 입력 컬럼 강조
    border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF'),
    )

    # 안내문 (1행)
    ws.cell(1, 1, '※ 가이드라인 시트 참조 후 [최종라벨] 컬럼에 P / N / X 입력. 변경 사유는 [메모] 컬럼에 한 줄 기록.')
    ws.cell(1, 1).font = Font(name='맑은 고딕', size=10, italic=True, color='C00000')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))

    # 헤더 (2행)
    cols = df.columns.tolist()
    for j, col in enumerate(cols, 1):
        c = ws.cell(2, j, col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border

    # 데이터 (3행~)
    for i, row in enumerate(df.itertuples(index=False), 3):
        for j, val in enumerate(row, 1):
            cell = ws.cell(i, j, val if not pd.isna(val) else '')
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = border
            col_name = cols[j - 1]
            if col_name == '플래그':
                cell.fill = flag_fill
            elif col_name == '최종라벨':
                cell.fill = label_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # 컬럼 너비
    widths = {
        'sample_idx': 9, '라벨러': 8, 'review_id': 12, 'brand': 8,
        'rating': 7, 'content_clean': 60, '속성': 14,
        'v1_라벨': 8, '자동변환': 9, '트리거매칭': 10, '플래그': 32,
        '최종라벨': 10, '메모': 25, '우선순위': 8,
    }
    for j, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = widths.get(col, 14)

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = 'A3'

    # 최종라벨 P/N/X 드롭다운 검증
    if '최종라벨' in cols:
        col_idx = cols.index('최종라벨') + 1
        letter = get_column_letter(col_idx)
        dv = DataValidation(
            type='list', formula1='"P,N,X"', allow_blank=True,
            errorTitle='입력 오류', error='P / N / X 중 하나만 입력 가능합니다.',
        )
        ws.add_data_validation(dv)
        last_row = 2 + len(df)
        dv.add(f'{letter}3:{letter}{last_row}')


def _write_review_workbook(df: pd.DataFrame, out_path: str) -> None:
    """가이드라인 + 검토 2시트 워크북 생성"""
    wb = Workbook()
    # 기본 시트 제거
    default = wb.active
    wb.remove(default)
    _write_guideline_sheet(wb, sheet_name='가이드라인')
    _write_review_sheet(wb, df, sheet_name='검토')
    wb.save(out_path)


def split_review_by_labeler(
    df_review: pd.DataFrame,
    out_dir: str = './final_data',
    p2_spot_check: int = 50,
) -> dict:
    """
    검토 후보를 라벨러별 + 우선순위별로 분할 저장.

    각 xlsx는 [가이드라인] + [검토] 2시트 구조로 생성.
    - 가이드라인: absa_guideline_v2.md 전문 렌더링
    - 검토: 실제 검토 케이스 + 최종라벨 입력 컬럼 (P/N/X 드롭다운)
    """
    os.makedirs(out_dir, exist_ok=True)
    paths = {}

    df_review = df_review.copy()
    df_review['우선순위'] = df_review['플래그'].apply(_classify_priority)

    for labeler, label_key in [('라벨러1', 'L1'), ('라벨러2', 'L2')]:
        sub = df_review[df_review['라벨러'] == labeler].copy()

        # P1: 즉시 검토
        p1 = sub[sub['우선순위'] == 'P1'].sort_values(['sample_idx', '속성']).reset_index(drop=True)
        p1_path = os.path.join(out_dir, f'absa_relabel_P1_{label_key}.xlsx')
        _write_review_workbook(p1, p1_path)

        # P2: 랜덤 50건 스폿 체크
        p2_pool = sub[sub['우선순위'] == 'P2']
        p2 = p2_pool.sample(min(p2_spot_check, len(p2_pool)), random_state=260506).sort_values(['sample_idx', '속성']).reset_index(drop=True)
        p2_path = os.path.join(out_dir, f'absa_relabel_P2_spotcheck_{label_key}.xlsx')
        _write_review_workbook(p2, p2_path)

        paths[labeler] = {
            'P1': (p1_path, len(p1)),
            'P2': (p2_path, len(p2)),
            'P2_total_pool': len(p2_pool),
        }

    return paths


# ════════════════════════════════════════════════════════════
# 4. 통계 보고
# ════════════════════════════════════════════════════════════

def report_conversion(
    df_v1_raw: pd.DataFrame,
    df_v2: pd.DataFrame,
    df_review: pd.DataFrame,
) -> None:
    print('=' * 70)
    print('  v1 → v2 변환 결과 보고')
    print('=' * 70)

    print('\n[1] U 라벨 자동 변환')
    for asp in ASPECTS:
        u_count = (df_v1_raw[asp] == 'U').sum()
        print(f'  {asp:18s}: U {u_count}건 → X')

    print('\n[2] v1 → v2 분포 변화')
    for asp in ASPECTS:
        v1_dist = df_v1_raw[asp].value_counts().to_dict()
        v2_dist = df_v2[asp].value_counts().to_dict()
        print(f'  {asp:18s}')
        print(f'    v1: {v1_dist}')
        print(f'    v2: {v2_dist}')

    print('\n[3] 라벨러별 재검토 후보 건수')
    by_labeler = df_review.groupby('라벨러').size().to_dict()
    for k, v in by_labeler.items():
        print(f'  {k}: {v}건')

    print('\n[4] 플래그 유형별 분포')
    flag_types = df_review['플래그'].str.split(' \\| ').explode()
    flag_types = flag_types.str.split(' \\(').str[0]
    print(flag_types.value_counts().to_string())

    print('\n[5] 속성별 재검토 후보')
    by_asp = df_review.groupby('속성').size().sort_values(ascending=False)
    print(by_asp.to_string())


# ════════════════════════════════════════════════════════════
# 5. 마스터 함수
# ════════════════════════════════════════════════════════════

def run_conversion(
    done_path: str = './final_data/absa_golden_set_1000_done.xlsx',
    out_dir:   str = './final_data',
) -> dict:
    """전체 파이프라인 실행"""
    os.makedirs(out_dir, exist_ok=True)

    df_v1_raw = pd.read_excel(done_path, sheet_name='라벨링', header=1)
    for col in ASPECTS:
        df_v1_raw[col] = df_v1_raw[col].fillna('X').astype(str).str.upper().str.strip()
        df_v1_raw.loc[~df_v1_raw[col].isin({'P', 'N', 'U', 'X'}), col] = 'X'

    df_v2, df_review = convert_v1_to_v2(done_path)

    v2_path = os.path.join(out_dir, 'absa_golden_set_1000_v2.xlsx')
    df_v2.to_excel(v2_path, index=False)

    review_paths = split_review_by_labeler(df_review, out_dir)

    report_conversion(df_v1_raw, df_v2, df_review)

    print(f'\n저장 완료:')
    print(f'  v2 라벨 파일: {v2_path}')
    for labeler, info in review_paths.items():
        p1_path, p1_n = info['P1']
        p2_path, p2_n = info['P2']
        print(f'  {labeler}:')
        print(f'    [P1 즉시검토] {p1_path}  ({p1_n}건)')
        print(f'    [P2 스폿체크] {p2_path}  (랜덤 {p2_n}건 / 전체 {info["P2_total_pool"]}건)')

    return {
        'v2_path':       v2_path,
        'review_paths':  review_paths,
        'total_review':  len(df_review),
    }

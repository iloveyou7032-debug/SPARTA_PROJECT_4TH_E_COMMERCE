"""
extract_boundary_cases.py — Phase 2 GT 재라벨링 boundary case 추출
====================================================================

목적: v9 모델 974건 검증셋 추론 결과 vs 골든셋 비교 → 라벨러가 집중 검토할
      150건의 'boundary case' Excel 생성. 핏/사이즈·브랜드/헤리티지 우선.

설계 원칙:
  - 사용자 안 (단순 mismatch + Pred 노출) + 5가지 개선
    A. Priority Score (핏·브랜드 ×2 / P↔N ×3 / overlap +2)
    B. Error Type Stratified (12 bucket 균등)
    C. Hint 자동 생성 (정규식 패턴 진단)
    D. 정답지 별도 파일 (anchoring 완화)
    E. v9 추론 캐시 (1회 ~30분, 이후 즉시 재사용)

Usage:
    cd 송원우/absa_exaone
    uv run python extract_boundary_cases.py

산출:
    absa_v9_validation_predictions.parquet     — 974건 v9 추론 캐시
    absa_relabel_boundary_150.xlsx             — 라벨러 작업용 (가이드/라벨링)
    absa_relabel_boundary_150_answers.xlsx     — 정답지 (검증용, 비공개)
"""
from __future__ import annotations

import re
import sys
import time
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd

ABSA_DIR = Path(__file__).resolve().parent
if str(ABSA_DIR) not in sys.path:
    sys.path.insert(0, str(ABSA_DIR))

from absa_v9 import (
    ASPECTS, LABELS,
    load_golden_set, build_few_shot_examples, build_batch_examples,
    predict_one,
)

# ─────────────────────────────────────────────────────────────
# 설정
# ─────────────────────────────────────────────────────────────
GOLDEN_PATH    = ABSA_DIR / "absa_golden_set_1000_v23.xlsx"
PRED_CACHE     = ABSA_DIR / "absa_v9_validation_predictions.parquet"
OVERLAP_PATH   = ABSA_DIR.parent / "final_data" / "kappa_overlap_100.xlsx"
OUT_LABELING   = ABSA_DIR / "absa_relabel_boundary_150.xlsx"
OUT_ANSWERS    = ABSA_DIR / "absa_relabel_boundary_150_answers.xlsx"

TARGET_N = 150
SEED = 42
TARGET_ASPECTS = ["핏/사이즈", "브랜드/헤리티지"]

# Error type 별 quota (합 150) — 라벨러 P recall 회복 + X 과추출 균형
BUCKET_QUOTA = {
    # 핏/사이즈 (75건)
    "핏/사이즈_X→P": 25,  # v9 최대 약점 (gold X 226건 → pred P)
    "핏/사이즈_X→N": 10,
    "핏/사이즈_P→X": 15,
    "핏/사이즈_N→X": 10,
    "핏/사이즈_P→N":  8,
    "핏/사이즈_N→P":  7,
    # 브랜드/헤리티지 (75건)
    "브랜드/헤리티지_X→P": 20,
    "브랜드/헤리티지_X→N":  5,
    "브랜드/헤리티지_P→X": 25,  # P recall 0.51 → 핵심 회복 영역
    "브랜드/헤리티지_N→X": 10,
    "브랜드/헤리티지_P→N":  8,
    "브랜드/헤리티지_N→P":  7,
}
assert sum(BUCKET_QUOTA.values()) == TARGET_N


# ─────────────────────────────────────────────────────────────
# Phase 0 — v9 추론 (캐시 또는 즉석 실행)
# ─────────────────────────────────────────────────────────────
def get_v9_predictions(golden: pd.DataFrame) -> pd.DataFrame:
    """검증셋 974건에 대한 v9 추론 결과 로드/생성. 캐시 1회 ~30분."""
    if PRED_CACHE.exists():
        print(f"[load] cache: {PRED_CACHE.name}")
        return pd.read_parquet(PRED_CACHE)

    fs = build_few_shot_examples(golden, n_per_class=3, seed=SEED)
    fs_contents = {ex_c for fs_list in fs.values() for ex_c, _ in fs_list}
    val = golden[~golden["content_clean"].isin(fs_contents)].reset_index(drop=True)
    print(f"[infer] v9 검증셋 추론 시작 ({len(val):,}건, M4 Pro 약 30분)")

    batch = build_batch_examples(golden, n_examples=6, seed=SEED)

    try:
        from tqdm import tqdm
        iterator = tqdm(val.iterrows(), total=len(val), desc="v9 추론")
    except ImportError:
        iterator = val.iterrows()

    preds = []
    start = time.time()
    for _, r in iterator:
        content = str(r["content_clean"])
        tokens  = str(r.get("tokens", ""))
        preds.append(predict_one(content, tokens, batch))
    elapsed = time.time() - start

    out = val[["sample_idx", "review_id", "content_clean"]].copy()
    for asp in ASPECTS:
        out[f"{asp}_pred"] = [p[asp] for p in preds]
    out.to_parquet(PRED_CACHE, index=False)
    print(f"[save] {PRED_CACHE.name} ({elapsed:.0f}s, {elapsed/len(val):.2f}s/건)")
    return out


# ─────────────────────────────────────────────────────────────
# Phase 1 — Priority Score
# ─────────────────────────────────────────────────────────────
def compute_priority_score(row: pd.Series) -> int:
    """
    가중치:
      - 타겟 속성(핏/브랜드) mismatch    ×2
      - 그 외 속성 mismatch              ×1
      - X↔P/N 혼동                       +1
      - P↔N 직접 혼동 (sentiment 반대)   +3
    """
    score = 0
    for asp in ASPECTS:
        g = row[asp]
        p = row[f"{asp}_pred"]
        if g == p:
            continue
        w = 2 if asp in TARGET_ASPECTS else 1
        if {g, p} == {"P", "N"}:
            score += w * 3
        elif "X" in (g, p):
            score += w * 1
    return score


def classify_error_type(gold: str, pred: str, aspect: str) -> Optional[str]:
    """타겟 속성에서 오류 타입 분류 (예: '핏/사이즈_X→P')."""
    if gold == pred:
        return None
    return f"{aspect}_{gold}→{pred}"


# ─────────────────────────────────────────────────────────────
# Phase 2 — Hint 자동 생성
# ─────────────────────────────────────────────────────────────
HINT_PATTERNS = {
    "편하다 단독":      re.compile(r"(편해|편하다|편한|편함|편할)"),
    "좋다·예쁘다·찰떡":  re.compile(r"(좋아요|좋다|좋네요|예뻐|예쁘다|이뻐|이쁘다|찰떡|딱이에요|가볍)"),
    "재구매·또사·강추":  re.compile(r"(또\s?사|또\s?살|재구매|재주문|다시\s?사|강추|단골|믿고\s?산|믿고\s?사)"),
    "사이즈 수치":       re.compile(r"\d+\s?(cm|mm|kg|호|사이즈)"),
    "핏 직접 단어":      re.compile(r"(핏|기장|허리|어깨|라인|와이존|Y존|압박|라이즈|밴드)"),
    "혼재 (-는데/-지만)": re.compile(r"(는데|지만|그러나|근데|하지만)"),
    "브랜드명 직접":     re.compile(r"(휠라|FILA|fila|안다르|젝시|룰루|레몬)"),
    "부정 톤":          re.compile(r"(별로|실망|환불|반품|비추|아쉬워|후회|짜증)"),
    "추천 표현":        re.compile(r"(추천|강추)"),
}


def make_hint(content: str, target_asp: str, gold: str, pred: str) -> str:
    """오류 패턴 자동 진단 — 라벨러 빠른 판단 보조."""
    hits = {name for name, pat in HINT_PATTERNS.items() if pat.search(content or "")}

    if target_asp == "핏/사이즈":
        if pred == "P" and gold == "X":
            if ("편하다 단독" in hits or "좋다·예쁘다·찰떡" in hits) and not (
                "사이즈 수치" in hits or "핏 직접 단어" in hits
            ):
                return "단순 호평(편하다·좋다 단독) — 핏 직접 언급 없으면 X 의심"
            return "핏 키워드 모호 → 사이즈/기장/허리 명시 여부 확인"
        if pred == "X" and gold in ("P", "N"):
            if "핏 직접 단어" in hits or "사이즈 수치" in hits:
                return "핏 단어 명시됨 → 모델이 X로 도망 가능성"
            return "간접 핏 표현(타이트·헐렁·붙는 느낌 등) 검토"
        if {gold, pred} == {"P", "N"} and "혼재 (-는데/-지만)" in hits:
            return "긍정+부정 혼재 — 후미 절 우선 규칙 적용"
        return "핏 mismatch 검토"

    if target_asp == "브랜드/헤리티지":
        if pred == "X" and gold == "P":
            if "재구매·또사·강추" in hits:
                return "재구매·강추 표현 명시 → 모델 누락 (P recall 약점)"
            if "브랜드명 직접" in hits and not ("부정 톤" in hits):
                return "브랜드명 명시 + 긍정 → P 누락 의심"
            if "추천 표현" in hits:
                return "간접 추천 시그널 → 모델이 X로 도망"
            return "간접 충성도 표현(하나더·평생·계속) 검토"
        if pred == "P" and gold == "X":
            if not ("재구매·또사·강추" in hits or "브랜드명 직접" in hits):
                return "단순 상품 칭찬을 브랜드 P로 오분류 (anchoring 의심)"
            return "브랜드 충성 강도 충분한지 재판단"
        if pred == "N" and gold == "X":
            return "단순 상품 불만을 브랜드 N으로 오분류 의심"
        if {gold, pred} == {"P", "N"}:
            return "브랜드 sentiment 정반대 — 부정 톤 재확인"
        return "브랜드 mismatch 검토"

    return ""


# ─────────────────────────────────────────────────────────────
# Phase 3 — Error Type Stratified Sampling
# ─────────────────────────────────────────────────────────────
def stratified_sample(merged: pd.DataFrame, seed: int = SEED) -> pd.DataFrame:
    """error type bucket 별 균등 추출 + priority_score 상위 우선."""
    rng = np.random.default_rng(seed)
    selected = set()
    rows = []
    shortfall = 0

    for asp in TARGET_ASPECTS:
        # 해당 속성 mismatch 후보군
        mismatch_mask = merged[asp] != merged[f"{asp}_pred"]
        cand_asp = merged[mismatch_mask].copy()
        cand_asp["err_type"] = cand_asp.apply(
            lambda r: classify_error_type(r[asp], r[f"{asp}_pred"], asp), axis=1
        )

        for err_type, quota in BUCKET_QUOTA.items():
            if not err_type.startswith(asp):
                continue
            bucket = cand_asp[
                (cand_asp["err_type"] == err_type)
                & (~cand_asp["sample_idx"].isin(selected))
            ].copy()

            if len(bucket) == 0:
                shortfall += quota
                print(f"  [skip] {err_type:>30s}  cand=0  quota={quota}")
                continue

            # priority_score 내림차순 → 상위 2*quota 풀에서 random 추출 (편향 완화)
            bucket = bucket.sort_values("priority_score", ascending=False)
            pool = bucket.head(quota * 2) if len(bucket) >= quota * 2 else bucket
            take = min(quota, len(pool))
            picked = pool.sample(n=take, random_state=seed) if take < len(pool) else pool

            print(f"  [pick] {err_type:>30s}  cand={len(bucket):>4}  pick={take:>3}/{quota}")

            for _, r in picked.iterrows():
                rd = r.to_dict()
                rd["target_aspect"] = asp
                rd["error_type"] = err_type
                rd["hint"] = make_hint(rd["content_clean"], asp, rd[asp], rd[f"{asp}_pred"])
                rows.append(rd)
                selected.add(rd["sample_idx"])

    if shortfall > 0:
        # 미달분은 priority_score 상위 잔여로 보충
        remaining = merged[
            ~merged["sample_idx"].isin(selected)
            & (
                (merged["핏/사이즈"] != merged["핏/사이즈_pred"])
                | (merged["브랜드/헤리티지"] != merged["브랜드/헤리티지_pred"])
            )
        ].sort_values("priority_score", ascending=False).head(shortfall)
        print(f"  [fill] shortfall={shortfall} 보충 ({len(remaining)}건)")
        for _, r in remaining.iterrows():
            rd = r.to_dict()
            # 어느 속성이 더 우선인지로 결정
            for asp in TARGET_ASPECTS:
                if rd[asp] != rd[f"{asp}_pred"]:
                    rd["target_aspect"] = asp
                    rd["error_type"] = classify_error_type(rd[asp], rd[f"{asp}_pred"], asp)
                    rd["hint"] = make_hint(rd["content_clean"], asp, rd[asp], rd[f"{asp}_pred"])
                    break
            rows.append(rd)
            selected.add(rd["sample_idx"])

    return pd.DataFrame(rows)


# ─────────────────────────────────────────────────────────────
# Phase 4 — Excel 출력
# ─────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────
# 가이드라인 본문 (라벨러 작업 표준)
# ─────────────────────────────────────────────────────────────
GUIDE_INTRO = [
    ("작업 목적",   "v9 모델(Macro F1=0.7032)이 헷갈린 boundary 150건을 라벨러1이 재검토 → "
                   "골든셋 v2.4 업데이트 → F1 0.78~0.81 도달 목표"),
    ("총 건수",     "150건 (핏/사이즈 75 + 브랜드/헤리티지 75)"),
    ("입력 라벨",    "P=긍정 언급 / N=부정 언급 / X=해당 속성 미언급"),
    ("작업 컬럼",    "_NEW × 6속성 (노란색 배경) — 6속성 모두 채워주세요"),
    ("작업 방식",    "_GOLD(현 라벨) / _PRED(v9 예측) 둘 다 참조 → 자기 판단으로 _NEW 결정"),
    ("예상 소요",    "건당 약 1~2분 → 총 150~300분 (분할 작업 권장)"),
    ("우선 검토",    "target_aspect 컬럼이 가리키는 속성을 가장 신중히 / 그 외 5속성도 일관성 위해 작성"),
    ("저장 시점",    "30~50건마다 Excel 저장 (작업 손실 방지)"),
    ("질의 채널",    "송원우에게 보더리 케이스 사진/문장 공유 → 같이 합의"),
]

GUIDE_ASPECTS = [
    {"속성": "핏/사이즈",
     "정의": "사이즈 수치(cm/kg), 기장, 허리, 어깨, 라인, Y존, 압박감, 라이즈, 정사이즈",
     "P 예시": "정사이즈 딱 맞음 / 라인 살아남 / Y존 커버",
     "N 예시": "사이즈 작음 / 허리 말림 / 답답함 / 한 치수 큰 듯",
     "X (미언급)": "편하고 좋네요 (단독) / 가볍다 (단독) / 색깔 예쁨"},
    {"속성": "소재/내구성",
     "정의": "원단, 촉감, 두께, 보풀, 변색, 비침, 늘어남, 세탁 후 변형",
     "P 예시": "원단 부드러움 / 두께 적당 / 탄탄함 / 변색 없음",
     "N 예시": "보풀 심함 / 비침 / 세탁 후 늘어남 / 까칠함",
     "X (미언급)": "디자인 예쁨 / 사이즈 딱 맞음 / 가성비 좋음"},
    {"속성": "기능성",
     "정의": "신축성, 통기성, 흡습/땀, 운동 활동성, 보온/냉감, 안 비침(기능)",
     "P 예시": "신축성 좋아 운동 편함 / 통기성 좋음 / 땀 안 참",
     "N 예시": "땀 차고 답답 / 안 늘어남 / 운동할 때 불편",
     "X (미언급)": "디자인 예쁨 / 핏 좋음 / 색감 마음에 듦"},
    {"속성": "디자인",
     "정의": "색상, 패턴, 실루엣, 외관, 코디 어울림, 시즌/계절감",
     "P 예시": "색감 예쁨 / 디자인 세련됨 / 코디 잘 어울림 / 실물 더 예쁨",
     "N 예시": "사진과 색상 다름 / 촌스러움 / 안 어울림 / 너무 화려",
     "X (미언급)": "사이즈 좋음 / 운동할 때 편함 / 가성비 좋음"},
    {"속성": "브랜드/헤리티지",
     "정의": "재구매 의사, 브랜드 충성·추천·강추·단골, 브랜드명(휠라/안다르/젝시믹스/룰루레몬) 직접 언급",
     "P 예시": "또 살게요 / 계속 쓸 듯 / 강추 / 단골 / 휠라 디자인 좋네요",
     "N 예시": "다시는 안 산다 / 실망 / 환불 / 비추 / 브랜드 가치 없음",
     "X (미언급)": "단순 상품 칭찬 (예쁘다·편하다 단독) / 디자인 좋음 (단독)"},
    {"속성": "가격/가치",
     "정의": "가격 만족/불만, 가성비, 합리적 가격, 할인, 비쌈, 저렴",
     "P 예시": "가성비 좋음 / 이 가격에 합리적 / 할인 좋음",
     "N 예시": "너무 비쌈 / 가격 거품 / 가격 대비 별로",
     "X (미언급)": "단순 가격 수치 언급 (6만원) / 배송비 불만 / 사이즈 좋음"},
]

GUIDE_RULES = [
    ("R1", "단순 호평 단독 = X",
     "'편해요·좋아요·찰떡·딱이에요·가볍다' **단독** 표현은 모든 속성 X. 핏/브랜드 직접 언급 없으면 X 유지"),
    ("R2", "디자인 평가어 단독 = 디자인 P 만",
     "'예쁘다·이쁘다·고급지다·세련' 단독은 디자인=P, 다른 속성=X (브랜드 P 아님)"),
    ("R3", "재구매·강추 = 브랜드 P",
     "'또 살게요·계속 쓸 듯·재구매·추천·강추·단골·믿고 산다' → 브랜드/헤리티지=P (브랜드명 없어도 OK)"),
    ("R4", "단순 상품 칭찬 ≠ 브랜드 P",
     "'예쁘다·편하다' 단독은 브랜드/헤리티지=X. 재구매·충성도 표현이 명시되어야 P"),
    ("R5", "혼재 시 후미 절 우선",
     "'~좋긴 한데 ~아쉽다' 같이 긍정+부정 혼재 시 끝부분 절을 따라 P/N 결정"),
    ("R6", "옵션 단어 = X",
     "'크롭·양면·셋업·반팔·긴팔' 등은 옵션 설명일 뿐 평가가 아님 → 해당 속성 X"),
    ("R7", "가격 수치만 = X",
     "'6만원·할인 받음' 같은 수치 단독은 가격=X. '이 가격에 가성비'만 가격=P"),
    ("R8", "신발/양말 착용감 = 기능성",
     "신발의 쿠셔닝·폭신함은 핏 아닌 기능성. 의류 착용감은 핏/사이즈"),
    ("R9", "끈/밴드 강도 = 소재",
     "끈·밴드의 늘어남·내구성은 기능 아닌 소재/내구성"),
]

GUIDE_ERROR_TYPES = [
    ("핏/사이즈_X→P", "정답 X인데 모델이 P로 오분류 (v9 최대 약점, 226건)",
     "본문에 사이즈/기장/허리 직접 언급 없으면 X 확정. '편하다·좋다 단독'이면 X 유지."),
    ("핏/사이즈_X→N", "정답 X인데 모델이 N으로 오분류",
     "본문에 핏 부정 표현(작다·말림 등) 없으면 X. 단순 불만은 핏 X."),
    ("핏/사이즈_P→X", "정답 P인데 모델이 X로 누락",
     "사이즈/핏 긍정 표현이 명확한지 재확인. 명확하면 P 유지, 애매하면 X로 수정."),
    ("핏/사이즈_N→X", "정답 N인데 모델이 X로 누락",
     "핏 부정 표현(작다·답답·말림)이 명확한지 재확인. 명확하면 N 유지."),
    ("핏/사이즈_P→N", "정답 P인데 모델이 N으로 잘못 판단",
     "긍정+부정 혼재 시 후미 절 우선 — 끝이 긍정이면 P, 부정이면 N."),
    ("핏/사이즈_N→P", "정답 N인데 모델이 P로 잘못 판단",
     "혼재 케이스에서 모델이 앞부분 긍정에 낚인 것. 끝부분 부정이면 N 유지."),
    ("브랜드/헤리티지_X→P", "정답 X인데 모델이 P로 오분류 (157건)",
     "단순 상품 칭찬을 브랜드 P로 오분류한 것. 재구매·강추 표현 없으면 X 확정."),
    ("브랜드/헤리티지_X→N", "정답 X인데 모델이 N으로 오분류",
     "단순 상품 불만을 브랜드 N으로 오분류한 것. '실망·다시는 안 산다' 등 직접 표현 없으면 X."),
    ("브랜드/헤리티지_P→X", "정답 P인데 모델이 X로 누락 (P recall 0.51, 핵심 회복 영역)",
     "재구매·강추·또 살게요·하나더 표현 있으면 P 확신. 모델이 간접 표현을 놓침."),
    ("브랜드/헤리티지_N→X", "정답 N인데 모델이 X로 누락",
     "실망·환불·비추 직접 표현 있으면 N 확신. 단순 불만이면 X."),
    ("브랜드/헤리티지_P→N", "정답 P인데 모델이 N으로 잘못 판단 (드문 case)",
     "전반 톤 재확인. 긍정 표현이 명확하면 P."),
    ("브랜드/헤리티지_N→P", "정답 N인데 모델이 P로 잘못 판단 (드문 case)",
     "부정 톤(별로·실망·환불) 명시되어 있으면 N. 모델이 앞부분 긍정에 낚인 것."),
]

GUIDE_CHECKLIST = [
    (1, "행 진입 시 가장 먼저 → target_aspect 와 error_type 확인 (어디를 신중히 봐야 할지)"),
    (2, "hint 컬럼 읽기 → v9가 왜 틀렸을지 추정 원인 파악 (절대 정답 아님, 판단 보조)"),
    (3, "content_clean 본문 정독 → 직접 언급 단어를 마음속으로 체크 (핏/소재/기능/디자인/브랜드/가격)"),
    (4, "_GOLD 와 _PRED 비교 → 누가 맞는지 판단 (둘 다 틀릴 수도 있음)"),
    (5, "_NEW 6속성 모두 작성 → P/N/X 중 하나. _GOLD에 동의하면 그대로, 변경 필요하면 새 라벨"),
    (6, "30~50건마다 저장 / 헷갈리면 송원우에게 사진 공유"),
]


def write_labeling_excel(sample_df: pd.DataFrame, out_path: Path) -> None:
    """라벨러 작업용 — 가이드 + 라벨링 + 분포 시트.

    가이드 시트는 5개 표를 한 시트에 순차 배치:
      0) 작업 개요 (intro)
      1) 6속성 정의 + P/N/X 예시
      2) 라벨링 결정 규칙 9개 (R1~R9)
      3) Error type 12개 해설 + 검토 포인트
      4) 작업 체크리스트 6단계
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    # 작업 시트 컬럼 순서
    cols_basic = ["sample_idx", "review_id", "brand", "rating",
                  "target_aspect", "error_type", "hint", "content_clean"]
    cols_new   = [f"{a}_NEW" for a in ASPECTS]

    out = pd.DataFrame()
    for c in cols_basic:
        out[c] = sample_df[c] if c in sample_df.columns else ""
    for a in ASPECTS:
        out[f"{a}_GOLD"] = sample_df[a]
    for a in ASPECTS:
        out[f"{a}_PRED"] = sample_df[f"{a}_pred"]
    for a in ASPECTS:
        out[f"{a}_NEW"]  = ""  # 라벨러 입력란

    dist = sample_df["error_type"].value_counts().reset_index()
    dist.columns = ["error_type", "count"]

    with pd.ExcelWriter(out_path, engine="openpyxl") as ew:
        # ── 가이드 시트 — 5개 표 순차 배치 ─────────────────
        ws_guide_name = "가이드"
        # 빈 DataFrame을 먼저 두고 시트 생성
        pd.DataFrame().to_excel(ew, sheet_name=ws_guide_name, index=False)
        ws = ew.sheets[ws_guide_name]

        # 스타일 정의
        title_font   = Font(bold=True, size=14, color="FFFFFFFF")
        title_fill   = PatternFill("solid", fgColor="FF1565C0")  # 진한 블루
        header_font  = Font(bold=True, size=11)
        header_fill  = PatternFill("solid", fgColor="FFE0E0E0")
        wrap_align   = Alignment(wrap_text=True, vertical="top")
        center_align = Alignment(horizontal="center", vertical="center")

        def write_section(title: str, rows: list, headers: list, start_row: int,
                          col_widths: list) -> int:
            """제목 + 표 작성 후 다음 시작 행 반환."""
            # 제목 행 (전체 컬럼 머지)
            ws.cell(row=start_row, column=1, value=title)
            ws.cell(row=start_row, column=1).font = title_font
            ws.cell(row=start_row, column=1).fill = title_fill
            ws.cell(row=start_row, column=1).alignment = center_align
            ws.row_dimensions[start_row].height = 22
            ws.merge_cells(start_row=start_row, start_column=1,
                           end_row=start_row, end_column=len(headers))

            # 헤더 행
            hdr_row = start_row + 1
            for c, h in enumerate(headers, start=1):
                cell = ws.cell(row=hdr_row, column=c, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = wrap_align

            # 데이터 행
            for i, row in enumerate(rows, start=hdr_row + 1):
                if isinstance(row, dict):
                    for c, h in enumerate(headers, start=1):
                        ws.cell(row=i, column=c, value=row.get(h, ""))
                        ws.cell(row=i, column=c).alignment = wrap_align
                else:
                    for c, val in enumerate(row, start=1):
                        ws.cell(row=i, column=c, value=val)
                        ws.cell(row=i, column=c).alignment = wrap_align
                ws.row_dimensions[i].height = 32

            # 컬럼 너비
            for c, w in enumerate(col_widths, start=1):
                ws.column_dimensions[get_column_letter(c)].width = w

            return start_row + 2 + len(rows) + 2  # 제목+헤더+데이터+빈2행

        # 0. 작업 개요
        nxt = write_section(
            "0.  작업 개요",
            GUIDE_INTRO,
            ["항목", "내용"],
            start_row=1,
            col_widths=[18, 80],
        )
        # 1. 6속성 정의
        nxt = write_section(
            "1.  6속성 정의 + P/N/X 예시",
            GUIDE_ASPECTS,
            ["속성", "정의", "P 예시", "N 예시", "X (미언급)"],
            start_row=nxt,
            col_widths=[18, 45, 35, 35, 35],
        )
        # 2. 라벨링 결정 규칙
        nxt = write_section(
            "2.  라벨링 결정 규칙 (R1~R9) — 헷갈릴 때 이 규칙 우선",
            GUIDE_RULES,
            ["코드", "규칙", "상세"],
            start_row=nxt,
            col_widths=[8, 26, 90],
        )
        # 3. Error type 해설
        nxt = write_section(
            "3.  Error Type 12종 — 라벨링 시트의 error_type 컬럼 해석",
            GUIDE_ERROR_TYPES,
            ["error_type", "의미", "검토 포인트"],
            start_row=nxt,
            col_widths=[24, 50, 70],
        )
        # 4. 작업 체크리스트
        nxt = write_section(
            "4.  작업 체크리스트 — 한 행 처리 시 이 순서대로",
            GUIDE_CHECKLIST,
            ["Step", "내용"],
            start_row=nxt,
            col_widths=[8, 100],
        )

        # ── 라벨링 시트 ────────────────────────────────────
        out.to_excel(ew, sheet_name="라벨링", index=False)
        ws_lab = ew.sheets["라벨링"]
        col_idx = {c: i + 1 for i, c in enumerate(out.columns)}
        ws_lab.column_dimensions[get_column_letter(col_idx["content_clean"])].width = 60
        ws_lab.column_dimensions[get_column_letter(col_idx["hint"])].width = 35
        ws_lab.column_dimensions[get_column_letter(col_idx["error_type"])].width = 22
        ws_lab.column_dimensions[get_column_letter(col_idx["target_aspect"])].width = 14
        for c in cols_new:
            ws_lab.column_dimensions[get_column_letter(col_idx[c])].width = 9
        for r in range(2, len(out) + 2):
            ws_lab.cell(row=r, column=col_idx["content_clean"]).alignment = wrap_align
            ws_lab.cell(row=r, column=col_idx["hint"]).alignment = wrap_align
            ws_lab.row_dimensions[r].height = 60

        # 헤더 강조 + _NEW 컬럼 노란색
        new_fill = PatternFill("solid", fgColor="FFFFF7CC")
        gold_fill = PatternFill("solid", fgColor="FFE3F2FD")
        pred_fill = PatternFill("solid", fgColor="FFFFEBEE")
        for c, idx in col_idx.items():
            cell = ws_lab.cell(row=1, column=idx)
            cell.font = header_font
            if c.endswith("_NEW"):
                cell.fill = new_fill
            elif c.endswith("_GOLD"):
                cell.fill = gold_fill
            elif c.endswith("_PRED"):
                cell.fill = pred_fill
            else:
                cell.fill = header_fill

        # 첫 행 고정 (스크롤해도 헤더 보임)
        ws_lab.freeze_panes = "I2"

        # ── 분포 시트 ─────────────────────────────────────
        dist.to_excel(ew, sheet_name="분포", index=False)


def write_answers_excel(sample_df: pd.DataFrame, out_path: Path) -> None:
    """정답지 별도 — 라벨러 작업 후 검증용 (anchoring 완화 목적)."""
    ans_cols = ["sample_idx", "review_id", "brand"]
    ans = sample_df[ans_cols].copy()
    for a in ASPECTS:
        ans[f"{a}_GOLD_v23"] = sample_df[a]
        ans[f"{a}_v9_PRED"] = sample_df[f"{a}_pred"]
    ans["error_type"] = sample_df["error_type"]
    ans["target_aspect"] = sample_df["target_aspect"]
    ans["priority_score"] = sample_df["priority_score"]
    ans.to_excel(out_path, index=False)


# ─────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────
def main() -> None:
    print("=" * 70)
    print("Phase 2 GT 재라벨링 — Boundary Case 추출")
    print("=" * 70)

    print("\n[1/4] 골든셋 + v9 predictions 로드")
    golden = load_golden_set(str(GOLDEN_PATH))
    pred = get_v9_predictions(golden)

    print("\n[2/4] Mismatch 분석 + Priority Score")
    merged = pred.merge(
        golden[["sample_idx", "brand", "rating"] + ASPECTS],
        on="sample_idx", how="inner",
    )
    merged["priority_score"] = merged.apply(compute_priority_score, axis=1)

    # 라벨러 overlap 가산
    if OVERLAP_PATH.exists():
        ov = pd.read_excel(OVERLAP_PATH)
        if "sample_idx" in ov.columns:
            ov_ids = set(ov["sample_idx"].tolist())
            mask = merged["sample_idx"].isin(ov_ids)
            merged.loc[mask, "priority_score"] += 2
            print(f"  라벨러 overlap {len(ov_ids)}건 → priority +2 가산: {mask.sum()}건 적중")
    else:
        print(f"  overlap 파일 없음 ({OVERLAP_PATH.name}) — overlap 가산 생략")

    has_target_mismatch = (
        (merged["핏/사이즈"] != merged["핏/사이즈_pred"])
        | (merged["브랜드/헤리티지"] != merged["브랜드/헤리티지_pred"])
    )
    cand = merged[has_target_mismatch].copy()
    print(f"  핏 또는 브랜드 mismatch: {len(cand):,}건 / 검증셋 {len(merged):,}건")

    print("\n[3/4] Stratified Sampling 150건")
    sampled = stratified_sample(cand, seed=SEED)
    print(f"\n  최종 추출: {len(sampled)}건")
    print("\n  [error_type 분포]")
    print(sampled["error_type"].value_counts().to_string())

    print("\n[4/4] Excel 출력")
    write_labeling_excel(sampled, OUT_LABELING)
    print(f"  ✓ {OUT_LABELING.name}  (가이드 + 라벨링 + 분포)")
    write_answers_excel(sampled, OUT_ANSWERS)
    print(f"  ✓ {OUT_ANSWERS.name}  (정답지 — 라벨러 작업 후 검증용)")
    print()
    print("=" * 70)
    print("완료. 라벨러에게 absa_relabel_boundary_150.xlsx 전달 후 _NEW 작성 요청.")
    print("=" * 70)


if __name__ == "__main__":
    main()

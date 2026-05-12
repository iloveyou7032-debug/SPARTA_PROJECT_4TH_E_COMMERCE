"""
extract_spotcheck_30.py — 라벨러1 _NEW 검증용 Spot-check 30건 추출
======================================================================

목적:
  라벨러1이 P/N → X 로 변경한 케이스 중 30건을 송원우가 독립 라벨링하여
  Cohen's κ 측정 → H1(v24가 옳다) vs H2(라벨러1 too strict) 판정.

배경:
  Phase 2 v10 추론 결과 Macro F1 0.6676 (Δ -0.0356, 시나리오 C).
  변경량 분석 결과 라벨러1은 압도적으로 P/N → X 보수화 경향
  (디자인 P→X 25건, 핏 N→X 17건 등). 송원우 독립 라벨링으로 검증 필요.

추출 대상 (4속성, 브랜드/가격 제외):
  핏/사이즈, 소재/내구성, 기능성, 디자인 — F1 하락 4속성

추출 비율 (변경량 비례, 30건):
  핏/사이즈:    7건 (P→X 3, N→X 4)   ← 전체 P→X 11, N→X 17
  소재/내구성:  7건 (P→X 4, N→X 3)   ← 전체 P→X 18, N→X 12
  기능성:       7건 (P→X 4, N→X 3)   ← 전체 P→X 14, N→X 13
  디자인:       9건 (P→X 6, N→X 3)   ← 전체 P→X 25, N→X 11

출력:
  absa_spotcheck_30_for_review.xlsx
   ├ "라벨링" 시트 — 30행 (송원우_LABEL 빈 칸)
   └ "가이드라인" 시트 — 4속성 정의 + 라벨링 원칙

다음 단계:
  송원우 작성 완료 → absa_spotcheck_30_completed.xlsx 저장
  → uv run python kappa_spotcheck.py

Usage:
  uv run python extract_spotcheck_30.py
"""
from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ABSA_DIR = Path(__file__).resolve().parent
if str(ABSA_DIR) not in sys.path:
    sys.path.insert(0, str(ABSA_DIR))

RELABEL_PATH = ABSA_DIR / "absa_relabel_boundary_150.xlsx"
OUT_PATH = ABSA_DIR / "absa_spotcheck_30_for_review.xlsx"
SEED = 42

EXTRACT_PLAN = {
    "핏/사이즈":   {"P→X": 3, "N→X": 4},
    "소재/내구성": {"P→X": 4, "N→X": 3},
    "기능성":      {"P→X": 4, "N→X": 3},
    "디자인":     {"P→X": 6, "N→X": 3},
}


# ─────────────────────────────────────────────────────────────
# 후보 추출
# ─────────────────────────────────────────────────────────────
def find_transitions(df: pd.DataFrame, aspect: str, from_label: str, to_label: str) -> pd.DataFrame:
    g = df[f"{aspect}_GOLD"].astype(str).str.upper()
    n = df[f"{aspect}_NEW"].astype(str).str.upper()
    return df[(g == from_label) & (n == to_label)].copy()


def build_spotcheck(relabel: pd.DataFrame, plan: dict, seed: int = SEED) -> pd.DataFrame:
    rows = []
    for asp, transitions in plan.items():
        for trans_str, n_req in transitions.items():
            from_l, to_l = trans_str.split("→")
            cands = find_transitions(relabel, asp, from_l, to_l)
            n_avail = len(cands)
            if n_avail < n_req:
                print(f"  [warn] {asp} {trans_str}: 후보 {n_avail}건 < 요청 {n_req} — 가능한 만큼 추출")
                n_req = n_avail
            sample = cands.sample(n=n_req, random_state=seed)
            for _, r in sample.iterrows():
                rows.append({
                    "spotcheck_id": "",
                    "sample_idx":  int(r["sample_idx"]),
                    "review_id":   int(r["review_id"]) if pd.notna(r["review_id"]) else "",
                    "brand":       r.get("brand", ""),
                    "rating":      int(r["rating"]) if pd.notna(r["rating"]) else "",
                    "content":     r["content_clean"],
                    "속성":         asp,
                    "v23_GOLD":    from_l,
                    "v24_라벨러1": to_l,
                    "송원우_LABEL": "",
                    "비고":         "",
                })
    df = pd.DataFrame(rows).sample(frac=1, random_state=seed).reset_index(drop=True)
    df["spotcheck_id"] = [f"SC-{i+1:03d}" for i in range(len(df))]
    return df


# ─────────────────────────────────────────────────────────────
# 스타일
# ─────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="4F81BD")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TARGET_FILL = PatternFill("solid", fgColor="FFF2CC")
ASPECT_FILL = PatternFill("solid", fgColor="D9E1F2")
GUIDE_TITLE_FILL = PatternFill("solid", fgColor="305496")
GUIDE_TITLE_FONT = Font(bold=True, color="FFFFFF", size=13)
SECTION_FILL = PatternFill("solid", fgColor="DDEBF7")
SECTION_FONT = Font(bold=True, size=11, color="1F4E79")
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _apply_section(ws, row, title, span=4):
    ws.cell(row=row, column=1, value=title)
    ws.cell(row=row, column=1).font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def _apply_text(ws, row, text, span=4):
    ws.cell(row=row, column=1, value=text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def write_excel(df: pd.DataFrame, path: Path) -> None:
    wb = Workbook()

    # ── Sheet 1: 라벨링 ──
    ws = wb.active
    ws.title = "라벨링"

    cols = list(df.columns)
    for j, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=j, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    label_col_idx = cols.index("송원우_LABEL") + 1
    aspect_col_idx = cols.index("속성") + 1

    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, v in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = BORDER
            if j == label_col_idx:
                cell.fill = TARGET_FILL
            if j == aspect_col_idx:
                cell.fill = ASPECT_FILL
                cell.font = Font(bold=True)

    widths = {
        "spotcheck_id": 12, "sample_idx": 10, "review_id": 12,
        "brand": 10, "rating": 8, "content": 60, "속성": 14,
        "v23_GOLD": 10, "v24_라벨러1": 12, "송원우_LABEL": 14, "비고": 30,
    }
    for col_name, w in widths.items():
        if col_name in cols:
            ws.column_dimensions[get_column_letter(cols.index(col_name)+1)].width = w

    ws.row_dimensions[1].height = 28
    for r in range(2, len(df)+2):
        ws.row_dimensions[r].height = 60

    ws.freeze_panes = "G2"

    # ── Sheet 2: 가이드라인 ──
    ws2 = wb.create_sheet("가이드라인")

    r = 1
    cell = ws2.cell(row=r, column=1, value="라벨링 일관성 검증 — Spot-check 30건")
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    cell.fill = GUIDE_TITLE_FILL
    cell.font = GUIDE_TITLE_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[r].height = 32
    r += 2

    # 작업 목적
    _apply_section(ws2, r, "▶ 작업 목적"); r += 1
    purpose = [
        "라벨러1이 P/N → X 로 변경한 케이스 중 30건을 송원우님이 독립 라벨링",
        "→ Cohen's κ 측정으로 라벨링 일관성 검증",
        "→ 결과에 따라 다음 단계 결정 (LoRA / 추가 라벨링 / 가이드 재정의)",
        "",
        "[배경] v10 추론 F1=0.6676 (Δ-0.0356) — 라벨러1 변경 196칸 중 보수화(P/N→X) 압도적",
        "       → 본 검증으로 \"v24가 옳음\" vs \"라벨러1 too strict\" 판정",
    ]
    for m in purpose:
        _apply_text(ws2, r, m); r += 1
    r += 1

    # 작성 요령
    _apply_section(ws2, r, "▶ 작성 요령"); r += 1
    rules = [
        "1. [송원우_LABEL] 컬럼에 P / N / X 중 하나 작성 (대소문자 무관)",
        "2. v23_GOLD / v24_라벨러1 라벨은 참고용 — 영향 받지 말고 독립 판단",
        "3. 애매한 경우 [비고] 컬럼에 짧게 (예: \"P/X 경계\", \"디자인 단서 약함\")",
        "4. 30건 모두 작성 → 같은 폴더에 absa_spotcheck_30_completed.xlsx 로 저장",
        "5. 다음: uv run python kappa_spotcheck.py",
    ]
    for m in rules:
        _apply_text(ws2, r, m); r += 1
    r += 1

    # 4속성 정의 표
    _apply_section(ws2, r, "▶ 4속성 정의 (검증 대상)"); r += 1
    table = [
        ["속성", "P (긍정)", "N (부정)", "X (언급 없음)"],
        ["핏/사이즈", "정사이즈, 예쁘게 핏, 기장 적당", "큼/작음, 안 맞음, 타이트, 기장 별로", "핏/사이즈 직접 언급 없음"],
        ["소재/내구성", "원단 좋음, 탄탄함, 오래 입어도 변형 없음", "보풀, 늘어남, 기모 빠짐, 까슬", "소재/내구성 직접 언급 없음"],
        ["기능성", "땀 안 참, 통기, 운동/요가 시 편함", "땀 참, 뻑뻑, 움직임 불편, 활동 제한", "기능 관련 언급 없음"],
        ["디자인", "예쁨, 색상 마음에 듦, 디자인 만족", "촌스러움, 색상 별로, 비율 이상", "디자인 직접 언급 없음 (\"이쁘다\"는 언급 인정)"],
    ]
    for i, trow in enumerate(table):
        for j, val in enumerate(trow, start=1):
            c = ws2.cell(row=r, column=j, value=val)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = BORDER
            if i == 0:
                c.fill = HEADER_FILL
                c.font = HEADER_FONT
                c.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[r].height = 36
        r += 1
    r += 1

    # 라벨링 원칙
    _apply_section(ws2, r, "▶ 핵심 라벨링 원칙 (라벨러1 가이드와 동일)"); r += 1
    principles = [
        "R1. 명시적 언급 우선 — 해당 속성을 텍스트에서 직접 언급한 경우만 P/N",
        "R2. 모호한 칭찬은 X — \"좋아요/만족\"만 있으면 어느 속성인지 모름 → X",
        "R3. 부정 단서 우선 — \"~라 아쉽\", \"~게 별로\" 등 부정 단서 있으면 N",
        "R4. 기능성 vs 디자인 분리 — \"편함\"=기능성 / \"이쁨\"=디자인",
        "R5. 강한 단서 1개 > 약한 단서 다수",
    ]
    for m in principles:
        _apply_text(ws2, r, m); r += 1
    r += 1

    # 변경 방향 가이드
    _apply_section(ws2, r, "▶ 추출된 30건의 변경 분포 (참고)"); r += 1
    dist = [
        "핏/사이즈:   7건 (P→X 3건, N→X 4건)",
        "소재/내구성: 7건 (P→X 4건, N→X 3건)",
        "기능성:      7건 (P→X 4건, N→X 3건)",
        "디자인:      9건 (P→X 6건, N→X 3건)",
        "총 30건 — 모두 \"v23이 P 또는 N으로 라벨, 라벨러1이 X로 재라벨\" 케이스",
    ]
    for m in dist:
        _apply_text(ws2, r, m); r += 1

    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 32
    ws2.column_dimensions['C'].width = 32
    ws2.column_dimensions['D'].width = 32

    wb.save(path)


# ─────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────
def main():
    print("=" * 70)
    print("[extract_spotcheck_30] Spot-check 30건 추출")
    print("=" * 70)

    if not RELABEL_PATH.exists():
        print(f"[abort] 입력 파일 없음: {RELABEL_PATH}")
        sys.exit(1)

    print(f"\n[1/3] 입력 로드: {RELABEL_PATH.name}")
    relabel = pd.read_excel(RELABEL_PATH, sheet_name="라벨링")
    print(f"  → {len(relabel)}건")

    print(f"\n[2/3] Stratified sampling (seed={SEED})")
    spotcheck = build_spotcheck(relabel, EXTRACT_PLAN, seed=SEED)
    print(f"  → 추출 {len(spotcheck)}건")

    print("\n  [추출 분포]")
    pivot = spotcheck.groupby(["속성", "v23_GOLD"]).size().unstack(fill_value=0)
    print(pivot.to_string())

    print(f"\n[3/3] Excel 저장")
    write_excel(spotcheck, OUT_PATH)
    print(f"  → {OUT_PATH}")

    print("\n" + "=" * 70)
    print("✓ 완료. 다음 단계:")
    print(f"  1. {OUT_PATH.name} 열기 → [송원우_LABEL] 컬럼 30건 작성")
    print(f"  2. 같은 폴더에 'absa_spotcheck_30_completed.xlsx' 로 저장")
    print(f"  3. uv run python kappa_spotcheck.py")
    print("=" * 70)


if __name__ == "__main__":
    main()
